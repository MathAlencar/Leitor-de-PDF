import tabula
import PyPDF2
import pandas as pd
import csv
import re
import os
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import numbers
import openpyxl
from PyPDF2 import PdfReader
from datetime import date
from pdfminer.high_level import extract_text
from decimal import Decimal

# Funções auxiliares para uso dentro da biblioteca.

# Capturando ano atual para realizar validações dentro de Extratos que não tem o ano na data principal.
data_atual = date.today()

def extraindo_texto(self, pdf, position, indice):
        try:
            reader = PyPDF2.PdfReader(pdf)
            page = reader.pages[indice]
            text = page.extract_text()

        except:
            text = extract_text(pdf)
        
        # alguns extratos estão com a data localizada em pontos diferentes.
        if position == 1:
            return text[:150]
        if position == 2:
            return text[600:]
        if position == 3:
            return text

def extraindo_texto_2(pdf, indice):
    reader = PyPDF2.PdfReader(pdf)
    page = reader.pages[indice - 1]
    
    return page.extract_text()
    
# alguns PDF tem erro de codificação, portanto essa função irá abrir cada um em determinada codigicação até conseguir encontrar a correta.
def decodificando_csv(lista, arquivo_csv):

    success = False

    encodings = ['utf-8', 'latin1', 'cp1252', 'iso-8859-15', 'ascii']

    for encoding in encodings:
        try:
            with open(arquivo_csv, 'r', encoding=encoding) as file:
                reader = csv.reader(file)
                for row in reader:
                    lista.append(row)
            success = True
            break
        except UnicodeDecodeError:
            print('continuando...')
        except Exception as e:
            print(f'Demais erro de codificação')  

    if not success:
        print('Não foi possível realizar a codificação deste CSV') 

def regex_padrao_data(texto, v_data):
    match v_data:
        case 1:
            data_pattern = r'\b\d{2}\s*/\s*[a-z]{3}\b'  # 02/set
        case 2:
            data_pattern = r"\d{2}/\d{2}/\d{4}"  # 01/08/2024

    return re.findall(data_pattern, texto)

def regex_padrao_valores(texto, v_valor):
    match v_valor:
        case 1:
            valor_pattern = r'-?\d{1,3}(?:\.\d{3})*,\d{2}' # -120,00 ou 50,00

    return re.findall(valor_pattern, texto)

def regex_padrao_texto(texto, v_texto):
    match v_texto:
        case 1:
            valor_texto = r".*[a-zA-Z]+.*" # -120,00 ou 50,00

    return re.findall(valor_texto, texto)

def resgatando_ano(data):
    data_atual_01 = date.today()

    mes = data[3:5]
                    
    mes_atual = data_atual_01.month

    if(int(mes) <= mes_atual):
        ano = data_atual_01.year
    else:
        ano = data_atual_01.year - 1

    return ano

def formantando_numeros(valor, v_formatacao):
        valores = valor
        if '-' in valores:
            valores = valores.replace('-', '')
            valores = valores.replace(".","").replace(",", '.')
            valores = float(f'-{valores}')
        else:
            valores = valores.replace(".","").replace(",", '.')
            valores = float(valores)

        return valores


class Estrutura_Padrao:
    
    def __init__(self, pdf, page_one, page_any, texto_padrao, arquivo_csv, lista, version_reader, showDFS):
        self._area_analisada = ''
        self.page_one = page_one
        self.page_any = page_any
        self.pdf = pdf
        self.texto_padrao = texto_padrao
        self.arquivo_csv = arquivo_csv
        self.lista = lista
        self.indice = 0
        self.version_reader = version_reader
        self.showDFS = showDFS
    
    def main_reader(self):
        num_pages = self.num_pages_pdf()

        for i in range(num_pages):
            self.indice+=1

            texto_padrao = self.texto_padrao

            text = self.extraindo_texto_2()

            if texto_padrao in text:
                self._area_analisada = self.page_one
            else:
                self._area_analisada = self.page_any

            match self.version_reader:
                case 1:
                    dfs = tabula.read_pdf(self.pdf, pages=self.indice, area=self._area_analisada) # analisar com área
                case 2:
                    dfs = tabula.read_pdf(self.pdf, pages=self.indice) # analisar sem área

            if self.showDFS is True:
                print(dfs)

            self.to_table(dfs)

        self.decodificando_csv()

        os.remove(self.arquivo_csv)

    def num_pages_pdf(self):
        try:
            with open(self.pdf, 'rb') as pdf_file:
                pdf_reader = PyPDF2.PdfReader(pdf_file)
                num_pages = len(pdf_reader.pages)
                return num_pages
        except:
            return print("número invalido")

    def extraindo_texto_2(self):
        reader = PyPDF2.PdfReader(self.pdf)
        page = reader.pages[self.indice - 1]
        
        return page.extract_text()
    
    def to_table(self, dfs):
        for table in dfs:
            table.to_csv(self.arquivo_csv, mode='a', index=False)

    def decodificando_csv(self):

        success = False

        encodings = ['utf-8', 'latin1', 'cp1252', 'iso-8859-15', 'ascii']

        for encoding in encodings:
            try:
                with open(self.arquivo_csv, 'r', encoding=encoding) as file:
                    reader = csv.reader(file)
                    for row in reader:
                        self.lista.append(row)
                success = True
                break
            except UnicodeDecodeError:
                print('continuando...')
            except Exception as e:
                print(f'Demais erro de codificação')  

        if not success:
            print('Não foi possível realizar a codificação deste CSV') 
    
    def regex_padrao_texto(texto, v_texto):
        match v_texto:
            case 1:
                valor_texto = r".*[a-zA-Z]+.*" 

        return re.findall(valor_texto, texto)

    def regex_padrao_data(texto, v_data):
        match v_data:
            case 1:
                data_pattern = r'\b\d{2}\s*/\s*[a-z]{3}\b'  # 02/set
            case 2:
                data_pattern = r"\d{2}/\d{2}/\d{4}"  # 01/08/2024

        return re.findall(data_pattern, texto)

    def regex_padrao_valores(texto, v_valor):
        match v_valor:
            case 1:
                valor_pattern = r'-?\d{1,3}(?:\.\d{3})*,\d{2}' # -120,00 ou 50,00
            case 2:
                valor_pattern = r"[+-]?\s*R\$\s*\d{1,3}(?:\.\d{3})*,\d{2}" # - R$ 1.712,67 ou - R$ 1.000,00
            case 3:
                valor_pattern = r"R\$ ?-?\d{1,3}(?:\.\d{3})*,\d+" # R$ -1.712,67 ou R$ -1.000,00
            case 4:
                valor_pattern = r"\d{1,3}(?:\.\d{3})*,\d{2} [DC]"
                
        return re.findall(valor_pattern, texto)

    def formantando_numeros(valor, v_formatacao):

        match v_formatacao:
            case 1:
                valores = valor
                if '-' in valores:
                    valores = valores.replace('-', '')
                    valores = valores.replace(".","").replace(",", '.')
                    valores = float(f'-{valores}')
                else:
                    valores = valores.replace(".","").replace(",", '.')
                    valores = float(valores)
            case 2:
                valores = valor
                if '-' in valores:
                    valores = valores.replace(' ', '').replace('-', '').replace('R$', '')
                    valores = float(f'-{valores.replace(".", "").replace(",", ".")}')
                else:
                    valores = valores.replace(' ', '').replace('+', '').replace('R$', '')
                    valores = float(f'{valores.replace(".", "").replace(",", ".")}')
            case 3:
                valores = valor
                if 'D' in valores:
                    valores = valores.replace(' ', '').replace('D', '')
                    valores = float(f'-{valores.replace(".", "").replace(",", ".")}')
                else:
                    valores = valores.replace(' ', '').replace('C', '')
                    valores = float(f'{valores.replace(".", "").replace(",", ".")}')  

        return valores

    def leitura_desc_separado(linha_1, linha_2, linha_3, v_leitura):

        descricao = ''

        match v_leitura:
            case 1:
                descricao = Estrutura_Padrao.leitura_desc_3_partes(linha_1, linha_2, linha_3)
                
        return descricao

    # versões de descrição.
    def leitura_desc_3_partes(linha_1, linha_2, linha_3):

        cont_atual_linha = 0
        cont_anterior_linha = 0
        cont_ante_anterior_linha = 0

        desc_now = ''
        desc_past = ''
        desc_past_past = ''

        for value in linha_1:
            if value != '':
                cont_atual_linha+=1
        
        for value in linha_2:
            if value != '':
                cont_anterior_linha+=1
        
        for value in linha_3:
            if value != '':
                cont_ante_anterior_linha+=1

        if cont_atual_linha == 1 and cont_ante_anterior_linha == 1:

            for value in linha_1:
                match_desc_ante_anterior = Estrutura_Padrao.regex_padrao_texto(value, 1)
                if match_desc_ante_anterior:
                    desc_now = match_desc_ante_anterior[0]
            
            for value in linha_2:
                match_desc_anterior = Estrutura_Padrao.regex_padrao_texto(value, 1)
                if match_desc_anterior:
                    descricao_anterior = match_desc_anterior[0]
            
            for value in linha_3:
                match_desc_atual = Estrutura_Padrao.regex_padrao_texto(value, 1)
                if match_desc_atual:
                    desc_past_past = match_desc_atual[0]

            juncao_desc = desc_past_past + ' ' + desc_now

            # print(' ')
            # print(linha_1)
            # print(linha_2)
            # print(juncao_desc)
            # print(' ')

            return juncao_desc

class Linha:
    def __init__(self, banco, data, descricao, valores, saldo_extrato, saldo_calculado, mes, ano, primeiro_dia_mes):
        self.banco = banco
        self.data = data
        self.descricao = descricao
        self.saldo_extrato = saldo_extrato
        self.saldo_calculado = saldo_calculado
        self.mes = mes
        self.ano = ano
        self.primeiro_dia_mes = primeiro_dia_mes

        pattern_letter = r"^[a-zA-Z\W]+$"
        validando_valores = f'{valores}'

        teste01 = re.match(pattern_letter, validando_valores)

        if validando_valores == '' or teste01:
            self.valores = 0
        else:
            self.valores = valores

class leitor_pdf_santander:
    
    def __init__(self):
        pass

    def lendo_pdf_santander_v1(self, pdf):

            data_pattern = r"\d{2}/\d{2}/\d{4}"
            texto = r"[a-zA-Z]\w*"

            fim_page = True

            def eh_valor_valido(valor):

                padrao = r"^-?\d{1,3}(?:\.\d{3})*,\d{2}$"
                match = re.match(padrao, valor)

                if valor == '':
                    return True

                return bool(match)

            try:
                with open(pdf, 'rb') as pdf_file:
                    pdf_reader = PyPDF2.PdfReader(pdf_file)
                    num_pages = len(pdf_reader.pages)
            except:
                return num_pages

            area_page01_satander = [[174,0,568,818]]
            area_restPages_santander = [[28,0,568,818]]
            lista_formatada = []
            area_analisada = area_page01_satander
            valores = 0

            indice_teste = 0

            for i in range(num_pages):
                credito = ''
                debito = ''
                saldo = ''

                indice = i+1

                dfs = tabula.read_pdf(pdf, pages=indice, area = area_analisada)
                lista = []

                for table in dfs:
                    table.to_csv("Santander_page.csv")

                    with open('Santander_page.csv', 'r') as file:
                        reader = csv.reader(file)
                        for row in reader:
                            lista.append(row)

                if fim_page == True:

                    for row in lista:
                        credito_validando, debito_validando, saldo_validando = row[-3:]

                        teste_01 = eh_valor_valido(credito_validando)
                        teste_02 = eh_valor_valido(debito_validando)
                        teste_03 = eh_valor_valido(saldo_validando) 

                        if teste_01 and teste_02 and teste_03:
                            credito, debito, saldo = row[-3:]

                        data =  ''
                        descricao = ''

                        for value in row:
                            match = re.match(data_pattern, value)
                            match_texto = re.match(texto, value)

                            teste_01 = eh_valor_valido(value)

                            if match:
                                data = value
                            if match_texto and teste_01 is False:
                                descricao = value

                        if data and descricao:
                                
                            if credito:
                                valores = credito
                            if debito:
                                valores = debito
                            

                            linha = Linha('Banco Santander', data, descricao, valores)
                            lista_formatada.append(linha)
                        
                        if descricao == '' and saldo:

                            if descricao == '' and credito == '' and debito == '' and saldo:
                                descricao = data[10:]
                                data = data[0:10]
                                
                                if credito:
                                    valores = credito
                                if debito:
                                    valores = debito

                                linha = Linha('Banco Santander', data, descricao, valores)
                                lista_formatada.append(linha)

                                indice_teste = indice + 2
                                fim_page = False

                                break

                            descricao = data[10:]
                            data = data[0:10]

                            if credito:
                                valores = credito
                            if debito:
                                valores = debito

                            linha = Linha('Banco Santander', data, descricao, valores)
                            lista_formatada.append(linha)

                        if data is False and descricao is False:
                            
                            descricao = ''
                            data = ''
                            
                            if credito:
                                valores = credito
                            if debito:
                                valores = debito
                            
                            linha = Linha('Banco Santander', data, descricao, valores)
                            lista_formatada.append(linha)

                area_analisada = area_restPages_santander

                if indice == indice_teste:
                    fim_page = True
                    area_analisada = area_page01_satander
            
            return lista_formatada

class leitor_pdf_bradesco:

    def __init__(self):
        pass

    def lendo_bradesco_celular_v1(self, pdf):
        #valores de dados
        primeiro_dia_mes = ''
        mes = ''
        ano = ''
        data = ''

        # armazenamento de dados
        lista_formatada = []
        lista_dados = []
        
        instanciando_classe = Estrutura_Padrao(pdf, [138,38,771,567], [0,0,0,0], 'Bradesco Celular', 'banco_bradesco_celular_csv', lista_dados, 1, False)

        instanciando_classe.main_reader()

        linha_anterior = ''
        linha_ante_anterior = ''

        for indice, row in enumerate(lista_dados):

            descricao = ''
            valores = 0
            cont_value = 0

            if indice >= 2:
                linha_anterior = lista_dados[indice - 1]
                linha_ante_anterior = lista_dados[indice - 2]

            descricao = Estrutura_Padrao.leitura_desc_separado(lista_dados[indice], linha_anterior, linha_ante_anterior, 1)

            for value in linha_anterior:
                match_data = Estrutura_Padrao.regex_padrao_data(value, 2)
                if match_data:
                    data = match_data[0]
                    mes = data[3:5]
                    ano = data[6:]
                    primeiro_dia_mes = f'01/{mes}/{ano}'
            
            for value in linha_anterior:
                match_valor = Estrutura_Padrao.regex_padrao_valores(value, 1)
                cont_value+=1
                if match_valor:
                    # realizando transoformação no número:
                    if cont_value >= 5:
                        match_valor[0] = f'-{match_valor[0]}'

                    valores = match_valor[0]
                    valores = Estrutura_Padrao.formantando_numeros(valores, 1)
                    break

            if valores != 0 and descricao:
                rowValue = Linha('banco Bradesco Celular - V1', data, descricao, valores,'', '', mes, ano, primeiro_dia_mes)
                lista_formatada.append(rowValue)

        return lista_formatada
    
    # Tentativa falha de leitura de PDF 
    def lendo_bradesco_simples_v2(self, pdf):
        #valores de dados
        primeiro_dia_mes = ''
        mes = ''
        ano = ''
        data = ''

        # armazenamento de dados
        lista_formatada = []
        lista_dados = []

        instanciando_classe = Estrutura_Padrao(pdf, [561, 37, 814, 125], [32, 32, 826, 555], 'Banco - bradesco Simples', 'banco_bradesco_csv', lista_dados, 1, False)

        instanciando_classe.main_reader()

        linha_anterior = ''
        linha_ante_anterior = ''

        for indice, row in enumerate(lista_dados):

            # print(row)

            valores = 0
            descricao = ''

            if indice >= 2:
                linha_anterior = lista_dados[indice - 1]
                linha_ante_anterior = lista_dados[indice - 2]

            descricao = Estrutura_Padrao.leitura_desc_separado(lista_dados[indice], linha_anterior, linha_ante_anterior, 1)

            for value in linha_anterior:
                buscando_data = Estrutura_Padrao.regex_padrao_data(value, 2)
                if buscando_data:
                    data = buscando_data[0]
                    mes = data[3:5]
                    ano = data[6:]
                    primeiro_dia_mes = f'01/{mes}/{ano}'
                    break

            for value in linha_anterior:
                match_valor = Estrutura_Padrao.regex_padrao_valores(value, 4)
                if match_valor:
                    valores = match_valor[0]
                    valores = Estrutura_Padrao.formantando_numeros(valores, 3)
                    break
            
            # print(descricao)

            if valores and descricao:
                # print('Banco - C6', data, descricao, valores,'','', mes, ano, primeiro_dia_mes)
                rowValue = Linha('Banco - C6', data, descricao, valores,'','', mes, ano, primeiro_dia_mes)
                lista_formatada.append(rowValue)

        return lista_formatada

class leitor_pdf_banco_itau:

    def __init__(self):
        pass

    def lendo_pdf_banco_itau_v1(self, pdf):

        # regex auxiliares
        data_pattern = r"\d{2}/\d{2}"
        texto_pattern = r".*[a-zA-Z]+.*"
        value_pattern = r"^\d{1,3}(?:\.\d{3})*,\d{2}-?$"
         
        # Código para pegar o número de páginas do PDF.
        with open(pdf, 'rb') as pdf_file:
            pdf_reader = PyPDF2.PdfReader(pdf_file)
            num_pages = len(pdf_reader.pages)
        
        # Extraindo a tabela de dados:

        area_analisada = [[121,17,757,560]]
        indice = 0
        lista = []
        lista_formatada = []

        # Variaveis de valores;
        mes = ''
        ano = ''
        primeiro_dia_mes = ''
        data = ''
        descricao = ''
        valores = ''

        for i in range(num_pages):

            indice+=1

            dfs = tabula.read_pdf(pdf, pages=indice, stream=True, area=area_analisada)

            for table in dfs:
                table.to_csv("Banco_itau_csv", mode='a', header=False, index=False)
            
        with open('Banco_itau_csv', 'r') as file:
                reader = csv.reader(file)
                for row in reader:
                    lista.append(row)    
        
        os.remove('Banco_itau_csv')

        for row in lista:
            
            # capturando data
            for value in row:
                validando_data = re.match(data_pattern, value)
                if validando_data:
                    
                    data = validando_data[0]
                    mes = data[3:5]

                    # Alguns extratos não tem o ano ao lado do mês e dia da transferência, então para ajustar isso apenas valido se o Mês é igual ou menor ao mês capturado, caso sim, se trata de um extrato de 2024, se não é um extrato do ano anterior.
                    mes_atual = data_atual.month

                    if(int(mes) <= mes_atual):
                        ano = data_atual.year
                    else:
                        ano = data_atual.year - 1

                    mes = data[3:5]
                    data = f'{validando_data[0]}/{ano}'

                    primeiro_dia_mes = f'01/{mes}/{ano}'
            
            for value in row:
                validando_texto = re.match(texto_pattern, value)
                if validando_texto:
                    descricao = value
                    break
            
            for value in row:
                validando_valor = re.match(value_pattern, value)
               
                if validando_valor:
                    valores_limpar = value

                    if '-' in valores_limpar:

                        valores_limpar = valores_limpar.replace('-', '')

                        valores = float(f'-{valores_limpar.replace(".", "").replace(",", ".")}') 
                        break
                    else:
                        valores = valores_limpar
                        valores = float(valores.replace(".", "").replace(",", "."))
                        break
            
            if row[-2]:
                print(valores)
                row_value = Linha('Banco Itau', data, descricao, valores, '','', mes, ano, primeiro_dia_mes)
                lista_formatada.append(row_value)

        return lista_formatada

    def leitor_pdf_itau_empresas_grafico(self, pdf):

        lista = []
        lista_formatada = []

        pageUm = [[354, 141, 821, 561]]
        AllPage = [[0, 82, 826, 563]]
        indice = 0

        area_analisada = AllPage

        with open(pdf, 'rb') as pdf_file:
            pdf_reader = PyPDF2.PdfReader(pdf_file)
            num_pages = len(pdf_reader.pages)
        
        for i in range(num_pages):
            indice+=1

            reader = PyPDF2.PdfReader(pdf)
            page = reader.pages[i]
            text = page.extract_text()

            flag_pagina_um = False
            flag_pagina_dois = False
            contador = 0

            if '01. Conta Corrente e Aplicações Automáticas' in text:
                area_analisada = pageUm
                flag_pagina_um = True
            
            lista_valido = ['data', 'descrição', 'entradas', 'saídas', 'saldo']

            for texto_valido in lista_valido:
                if texto_valido in text:
                    contador = contador + 1

            if contador == 5:
                flag_pagina_dois = True

            if 'Conta Corrente | Saques efetuados' in text or 'Conta Corrente | Débitos automáticos efetuados' in text or '02. Investimentos' in text or 'Notas explicativa' in text or '03. Crédito' in text or 'Débitos automáticos efetuados' in text or 'Pacote de serviços' in text:
                if flag_pagina_um is False and flag_pagina_dois is False:
                    continue
                else:
                    nada = ''

            if flag_pagina_um is False:
                area_analisada = AllPage
            
            dfs = tabula.read_pdf(pdf, pages=indice, area=area_analisada)

            for table in dfs:
                table.to_csv("Banco_bradesco_csv", mode='a', header=False, index=False)
            
        with open('Banco_bradesco_csv', 'r', encoding='utf-8') as file:
            reader = csv.reader(file)
            for row in reader:
                lista.append(row)

        os.remove("Banco_bradesco_csv")

        mes = ''
        ano = ''
        primeiro_dia_mes = ''
        data = ''
        descricao = ''
        valores = ''

        for row in lista:

            descricao = ''
            valores = ''
            valores_validando = '' # Usada para criar uma cópia do valor, e ve se ele contém letras.

            validando_data_final = bool(re.match(r"\d{2}/\d{2}/\d{2}", row[0]))

            validando_data_personnalite = r"\b\d{2}/\d{2}\b"
            
            for value in row:
                buscando_data = re.findall(validando_data_personnalite, value)
                if len(value) <= 6:
                    if buscando_data:
                        data = buscando_data[0]
                        mes = data[3:5]

                        # Alguns extratos não tem o ano ao lado do mês e dia da transferência, então para ajustar isso apenas valido se o Mês é igual ou menor ao mês capturado, caso sim, se trata de um extrato de 2024, se não é um extrato do ano anterior.
                        mes_atual = data_atual.month

                        if(int(mes) <= mes_atual):
                            ano = data_atual.year
                        else:
                            ano = data_atual.year - 1

                        mes = data[3:5]

                        data = f'{buscando_data[0]}/{ano}'

                        primeiro_dia_mes = f'01/{mes}/{ano}'
                        break
                else:
                    nada = ''

            # Validando data pois se caso os meses for acima de 12, se trata de uma informação invalida
            valid_mes = False
            if data != '':
                validando_mes = data[4:5]

                if int(validando_mes) > 12:
                    valid_mes = True

            # Caso as duas buscas por data for falso, aqui ele irá realizar uma nova busca dentro da Linha.
            
            for value in row:
                tem_letras_01 = bool(re.search(r"[a-zA-Z]", value))
                if tem_letras_01 and len(value) > 2:
                    descricao = value

            # Validando/encontrando valores positivos e negativos, onde os mesmo podem estar localizados nos indices do row.
            # validando_valor = r"\d+,\d{2}"

            validando_valor = r"^(?:\d{1,3}(?:\.\d{3})*,\d{2}(?:[-+])?)$"

            for value in row:
                match = re.match(validando_valor, value)

                if match:
                    valores = value
                    valores_validando = value
                    if '-' in valores:
                        valores = valores.replace('-', '')
                        valores = valores.replace(".","").replace(",", '.')
                        valores = float(f'-{valores}')
                    else:
                        valores_validando = value
                        valores = valores.replace(".","").replace(",", '.')
                        valores = float(valores)
                    break

            # Valida se o valor encontrado contém texto, caso sim, ele irá retornar True.

            tem_letras_valores = bool(re.search(r"[a-zA-Z]", valores_validando))

            descricao_invalida = ['na conta corrente (1)','principal',
            'l√≠quido','A Fechamento', 'bruto','do DI', 'líquido', 'Depósitos e recebimentos','DOCs e TEDs'
            'Outras entradas','Saques efetuados','Débitos automáticos efetuados','Outras saídas','Saldo anterior', 'Saldo final', 'Saldo anterior', 'Saldo em C/C']

            # Flag para verificar se a descrição é invalida
            desc_invalida = False

            for validando in descricao_invalida:
                if validando in descricao:
                    desc_invalida = True

            if descricao and valores and desc_invalida is False and valid_mes is False and data != '':
                # Caso validar a data inicial, assim evitando duplicatas, e tbm se há valores inválidos nos números, ele irá adicionar o valor no array.
                if validando_data_final is False and tem_letras_valores is False:
                    # na última verificação eu tiro os valores invalidos com base no mês, já que ele está pegando valores acima de 12.
                    if int(mes) <= 12:
                        rowValue = Linha('Banco Itau Empresa - Personnalite/PJ', data, descricao, valores, '', '',mes, ano, primeiro_dia_mes)
                        lista_formatada.append(rowValue)
                    else:
                        nada = ''

        return lista_formatada

    # Itau Uniclass
    def leitor_pdf_itau_uniclass(self, pdf):

        with open(pdf, 'rb') as pdf_file:
            pdf_reader = PyPDF2.PdfReader(pdf_file)
            num_pages = len(pdf_reader.pages)

        print(num_pages)
        
        areaPageUm = [[259, 23, 825, 572]]
        areaPageAll = [[19, 19, 828, 570]]

        # regex para buscar valores padrões:
        data_pattern = r"\d{2}/\d{2}/\d{4}"
        texto = r".*[a-zA-Z]+.*"
        validando_valor = r"^-?\d{1,3}(?:\.\d{3})*,\d{2}$"
        teste_valor = r"-?\d{1,3}(?:\.\d{3})*,\d{2}"

        # Variaveis de validação.
        data = ''
        descricao = ''
        indice = 0

        splitando = []

        lista = []
        lista_formatada = []

        area_analisada = ''
        
        for i in range(num_pages):
            indice+=1

            if indice == 1:
                area_analisada = areaPageUm
            else:
                area_analisada = areaPageAll
            
            dfs = tabula.read_pdf(pdf, pages=indice, area=area_analisada, stream=True)
            
            for table in dfs:
                table.to_csv('Banco_itau_extrato_simples.csv', mode='a', index=False)

        with open('Banco_itau_extrato_simples.csv', 'r', encoding='utf-8') as file:
            reader = csv.reader(file)
            for row in reader:
                lista.append(row)
        
        os.remove('Banco_itau_extrato_simples.csv')

        mes = ''
        ano = ''
        primeiro_dia_mes = ''
        data = ''
        descricao = ''
        
        for row in lista:

            valores = ''

            for value in row:
                match_data = re.match(data_pattern, value)
                match_valor = re.match(validando_valor, value)

                if match_data:
                    data = value[:10]
                    ano = data[6:]
                    mes = data[3:5]
                    primeiro_dia_mes = f'01/{mes}/{ano}'
                if match_valor:
                    valores_verificacao = value
                    
                    valores_verificacao = valores_verificacao.replace(".","").replace(",", ".")
                    valores = float(valores_verificacao)
                    break

                if len(row) >= 2:
                    splitando = row[1].split()

                # Separando o valor da descrição, já que em algumas colunas ambos vem grudados.
                for value in splitando:
                    match = re.search(validando_valor, value)
                    if match:
                        valores = match.group(0)
                
            for value in row:
                    match_descricao = re.match(texto, value)

                    if match_descricao:
                        
                        # Aqui você encontra uma gambiarra para solucionar a vinda de descrição com valores.
                        descricao = value
                        
                        valindaod_data_descricao = re.match(data_pattern, value)

                        if valindaod_data_descricao:
                            data_desc = valindaod_data_descricao[0]
                            descricao = descricao.replace(data_desc, '')
                        break

            # última tentativa de capturar o valor:
            if valores == '':
                for value in row:
                    match_valor = re.findall(teste_valor, value)
                    if len(match_valor) >= 1:
                        valores_verificacao = match_valor[0]
                        valores_verificacao = valores_verificacao.replace(".","").replace(",", ".")
                        valores = float(valores_verificacao)

            valores_desc_indesejados = [
                'SALDO FINAL',
                '(=) LIMITE DA CONTA DISPONÍVEL',
                '(+) LIMITE DA CONTA TOTAL',
                '(-) LIMITE DA CONTA UTILIZADO',
                '(=) LIMITE DA CONTA DISPONÕVEL',
                'saldo disponível sem investimentos automáticos',
                '(+) saldo em aplicação automática - aplic aut mais',
                '(+) rendimentos de aplicações automáticas',
                '(+) rendimentos de aplicações automáticas',
                 '(+) saldo em aplicação automática - poup aut',
                 '(=) saldo total disponível',
                 '(+) saldo bloqueado - dep. cheques',
                 '(+) saldo em aplicação automática - poup aut',
                 'LIMITE DA CONTA CONTRATADO',
                 'JUROS DO LIMITE DA CONTA ',
                 'IOF',
                 'detalhamento valor',
                 'TAXA EFETIVA MENSAL',
                 'CUSTO EFETIVO TOTAL (CET) MENSAL',
                 'CUSTO EFETIVO TOTAL (CET) ANUAL ',
                 'Unnamed: 0'
                ]

            flag_verificar_desc = False

            for desc in valores_desc_indesejados:
                if desc in descricao:
                    flag_verificar_desc = True
                    break

            if valores and descricao and flag_verificar_desc is False:
                rowValue = Linha('banco itaú - uniclass', data, descricao, valores,'', '', mes, ano, primeiro_dia_mes)
                lista_formatada.append(rowValue)

        return lista_formatada

    # ItaúEmpresas
    def leitor_pdf_itau_empresas(self, pdf):

        with open(pdf, 'rb') as pdf_file:
            pdf_reader = PyPDF2.PdfReader(pdf_file)
            num_pages = len(pdf_reader.pages)

        # regex para buscar valores padrões:
        data_pattern = r"\d{2}/\d{2}"
        texto = r".*[a-zA-Z]+.*"
        validando_valor = r"^-?\d{1,3}(?:\.\d{3})*,\d{2}$"
        teste_valor = r"-?\d{1,3}(?:\.\d{3})*,\d{2}"

        # Variaveis de validação.
        areaPageUm = [[150, 7, 836, 583]]
        areaPageAll = [[19, 19, 828, 570]]

        # variaveis auxiliares
        lista = []
        lista_formatada = []
        areaAnalisada = ''
        indice = 0
        nada = ''
        
        for i in range(num_pages):
            indice+=1

            if indice == 1:
                areaAnalisada = areaPageUm
            else:
                areaAnalisada = areaPageAll
                
            dfs = tabula.read_pdf(pdf, pages=indice, area=areaAnalisada)
                
            for table in dfs:
                table.to_csv('Banco_itau_extrato_empresas.csv', mode='a', index=False)

        with open('Banco_itau_extrato_empresas.csv', 'r', encoding='utf-8') as file:
            reader = csv.reader(file)
            for row in reader:
                lista.append(row)
        
        os.remove('Banco_itau_extrato_empresas.csv')

        # Variaveis de busca de valores.
        mes = ''
        ano = ''
        primeiro_dia_mes = ''
        data = ''
        descricao = ''

        for row in lista:
            
            valores = ''

            for value in row:
                match_data = re.match(data_pattern, value)
                match_valor = re.match(validando_valor, value)

                if match_data:
                    data = value
                    mes = data[3:5]
                    
                    mes_atual = data_atual.month

                    if(int(mes) <= mes_atual):
                        ano = data_atual.year
                    else:
                        ano = data_atual.year - 1

                    mes = data[3:5]

                    data = f'{data}/{ano}'

                    primeiro_dia_mes = f'01/{mes}/{ano}'

                if match_valor:
                    valores_verificacao = value
                    
                    valores_verificacao = valores_verificacao.replace(".","").replace(",", ".")
                    valores = float(valores_verificacao)
                    break
                
            for value in row:
                    match_descricao = re.match(texto, value)

                    if match_descricao:
                        
                        # Aqui você encontra uma gambiarra para solucionar a vinda de descrição com valores.
                        descricao = value
                        
                        valindaod_data_descricao = re.match(data_pattern, value)

                        if valindaod_data_descricao:
                            data_desc = valindaod_data_descricao[0]
                            descricao = descricao.replace(data_desc, '')
                        break

            if descricao and valores:
                if int(mes) <= 12:
                    rowValue = Linha('Banco Itau Empresas - itaú empresas', data, descricao, valores, '', '',mes, ano, primeiro_dia_mes)
                    lista_formatada.append(rowValue)
                else:
                    nada = ''
            
        return lista_formatada

    # ItaúEmpresas Simples
    def leitor_pdf_itau_empresasSimples(self, pdf):
        
        # variaveis de controle.
        num_pages = 0       
        indice = 0

        # Área das tabelas
        area_analisada = ''
        page_um = [233,31,820,656]
        page_any = [30,26,818,564]

        # regex dos valores capturados.
        data_pattern = r"\d{2}/\d{2}/\d{4}"
        texto = r".*[a-zA-Z]+.*"
        validando_valor = r"^-?\d{1,3}(?:\.\d{3})*,\d{2}$"
        teste_valor = r"-?\d{1,3}(?:\.\d{3})*,\d{2}"

        #valores de dados
        primeiro_dia_mes = ''
        dia = ''
        mes = ''
        ano = ''
        data = ''

        # Armazenar valores
        lista = []
        lista_formatada = []

        try:
            with open(pdf, 'rb') as pdf_file:
                pdf_reader = PyPDF2.PdfReader(pdf_file)
                num_pages = len(pdf_reader.pages)
        except:
            return print("número invalido")

        for i in range(num_pages):
            indice+=1

            texto_padrao = 'Saldo total Limite da conta Utilizado Disponível'

            text = extraindo_texto_2(pdf, indice)

            if texto_padrao in text:
                area_analisada = page_um
            else:
                area_analisada = page_any
            
            dfs = tabula.read_pdf(pdf, pages=indice, area=area_analisada)

            for table in dfs:
                table.to_csv("banco_itáu_EmpresasSimples_csv", mode='a', index=False)

        decodificando_csv(lista, 'banco_itáu_EmpresasSimples_csv')

        os.remove('banco_itáu_EmpresasSimples_csv')

        for row in lista:

            descricao = ''
            valores = 0

            for value in row:
                match_data = re.match(data_pattern, value)
                if match_data:
                    data = value[:10]
                    ano = data[6:]
                    mes = data[3:5]
                    primeiro_dia_mes = f'01/{mes}/{ano}'

            for value in row:
                match_valor = re.match(validando_valor, value)

                if match_valor:
                    valores = value
                    if '-' in valores:
                        valores = valores.replace('-', '')
                        valores = valores.replace(".","").replace(",", '.')
                        valores = float(f'-{valores}')
                    else:
                        valores = valores.replace(".","").replace(",", '.')
                        valores = float(valores)
                    break
                
            for value in row:
                match_descricao = re.match(texto, value)

                if match_descricao:
                    descricao+=' '+match_descricao[0]
            
            if valores and descricao:
                rowValue = Linha('banco itaú - EmpresasSimples ', data, descricao, valores,'', '', mes, ano, primeiro_dia_mes)
                lista_formatada.append(rowValue)
            
        return lista_formatada

    # ItaúEmpresas Simples 2
    def leitor_pdf_itau_empresasSimples2(self, pdf):
        #valores de dados
        primeiro_dia_mes = ''
        mes = ''
        ano = ''
        data = ''

        # armazenamento de dados
        lista_formatada = []
        lista_dados = []
        
        instanciando_classe = Estrutura_Padrao(pdf, [249,18,808,571], [27,21,816,571], 'limite da conta contratado', 'banco_itáu_EmpresasSimples2_csv', lista_dados, 1, False)

        instanciando_classe.main_reader()

        for row in lista_dados:

            descricao = ''
            valores = 0
            
            for value in row:
                match_data = regex_padrao_data(value, 1)
                if match_data:
                    data = match_data[0]
                    data = data.replace(' ', '').replace('jan', '01').replace('fev', '02').replace('mar', '03').replace('abri', '04').replace('mai', '05').replace('jun', '06').replace('jul', '07').replace('ago', '08').replace('set', '09').replace('out', '10').replace('nov', '11').replace('dez', '12')
                    ano = resgatando_ano(data)
                    mes = data[3:5]
                    data = f'{data}/{ano}'
                    primeiro_dia_mes = f'01/{mes}/{ano}'

            for value in row:
                match_valor = regex_padrao_valores(value, 1)
                if match_valor:
                    valores = match_valor[0]
                    valores = Estrutura_Padrao.formantando_numeros(valores, 1)
                    break
            
            for value in row:
                match_descricao = regex_padrao_texto(value, 1)
                if match_descricao:
                    if 'Unnamed' in match_descricao[0]:
                        continue
                    else:
                        descricao = match_descricao[0]
                        
            if valores and descricao:
                rowValue = Linha('banco itaú - EmpresasSimples2', data, descricao, valores,'', '', mes, ano, primeiro_dia_mes)
                lista_formatada.append(rowValue)
            
        return lista_formatada

    # itau Digital
    def leitor_pdf_itau_digital(self, pdf):
         #valores de dados
        primeiro_dia_mes = ''
        mes = ''
        ano = ''
        data = ''

        # armazenamento de dados
        lista_formatada = []
        lista_dados = []
        
        instanciando_classe = Estrutura_Padrao(pdf, [0, 0, 0, 0], [0, 0, 0, 0], 'dados geraisnome', 'banco_itáu_digital_csv', lista_dados, 2, False)

        instanciando_classe.main_reader()

        linha_anterior = lista_dados[0]

        for indice, row in enumerate(lista_dados):

            # linha anterior
            descricao_anterior = ''
            valores_anterior = ''
            descricao = ''
            valores = 0

            # flag para nao adicionar linhas repetidas
            flag_linha_repetida = False

            if indice >= 1:
                linha_anterior = lista_dados[indice - 1]

                for value in linha_anterior:
                    match_valor = Estrutura_Padrao.regex_padrao_valores(value, 2)
                    if match_valor:
                        valores_anterior = match_valor[0]
                        valores_anterior = Estrutura_Padrao.formantando_numeros(valores_anterior, 2)
                        break
                
                for value in linha_anterior:
                    match_descricao = Estrutura_Padrao.regex_padrao_texto(value, 1)
                    if match_descricao:
                        if 'Unnamed' in match_descricao[0]:
                            continue
                        else:
                            descricao_anterior = match_descricao[0]
                            break

            for value in lista_dados[indice]:
                match_data = Estrutura_Padrao.regex_padrao_data(value, 2)
                if match_data:
                    data = match_data[0]
                    mes = data[3:5]
                    ano = data[6:]
                    primeiro_dia_mes = f'01/{mes}/{ano}'
            
            for value in lista_dados[indice]:
                match_valor = Estrutura_Padrao.regex_padrao_valores(value, 2)
                if match_valor:
                    valores = match_valor[0]
                    valores = Estrutura_Padrao.formantando_numeros(valores, 2)
                    break
            
            for value in lista_dados[indice]:
                match_descricao = Estrutura_Padrao.regex_padrao_texto(value, 1)
                if match_descricao:
                    if 'Unnamed' in match_descricao[0]:
                        continue
                    else:
                        descricao = match_descricao[0]
                        break
            
            if descricao == descricao_anterior and valores_anterior == valores:
                flag_linha_repetida = True
                
            if valores != 0 and descricao and flag_linha_repetida is False:
                rowValue = Linha('banco itaú - Digital', data, descricao, valores,'', '', mes, ano, primeiro_dia_mes)
                lista_formatada.append(rowValue)

        for row in lista_dados:
            
            descricao = ''
            valores = 0

            # for value in row:
            #     match_data = Estrutura_Padrao.regex_padrao_data(value, 2)
            #     if match_data:
            #         data = match_data[0]
            #         mes = data[3:5]
            #         ano = ano = data[6:]
            #         data = f'{data}/{ano}'
            #         primeiro_dia_mes = f'01/{mes}/{ano}'

            # for value in row:
            #     match_valor = Estrutura_Padrao.regex_padrao_valores(value, 2)
            #     if match_valor:
            #         valores = match_valor[0]
            #         valores = Estrutura_Padrao.formantando_numeros(valores, 2)
            #         break
            
            # for value in row:
            #     match_descricao = Estrutura_Padrao.regex_padrao_texto(value, 1)
            #     if match_descricao:
            #         if 'Unnamed' in match_descricao[0]:
            #             continue
            #         else:
            #             descricao = match_descricao[0]
            #             break

            # if valores != 0 and descricao:
            #     rowValue = Linha('banco itaú - Digital', data, descricao, valores,'', '', mes, ano, primeiro_dia_mes)
            #     lista_formatada.append(rowValue)

        return lista_formatada

# Leitor PDF Banco do brasil;
class lendo_pdf_brasil:

    def __init__(self):
        pass
    
    def lendo_pdf_brasil_v1(self, pdf):

        try:
            with open(pdf, 'rb') as pdf_file:
                pdf_reader = PyPDF2.PdfReader(pdf_file)
                num_pages = len(pdf_reader.pages)
        except:
            return print("número invalido")

        lista_formatada = []

        # Variaveis de uso;
        indice = 0
        area_analisada = [[85,20,820,580]]

        # Regex para capturar dados:
        data_pattern = r'\d{2}/\d{2}/\d{4}'
        value_pattern = r'\d{1,3}(?:\.\d{3})*,\d{2} \(\+\)|\d{1,3}(?:\.\d{3})*,\d{2} \(-\)'
        retirando_desc = r'\d+,\d{2} \(\+\)|\d+,\d{2} \(-\)|\d{2}:\d{2}|\d{2}/\d{2}/\d{4}|\d{2}/\d{2}'

        for i in range(num_pages):
            indice+=1
            dfs = tabula.read_pdf(pdf, pages=indice,  lattice=True, area=area_analisada)

            lista = []

            for table in dfs:
                table.to_csv("Banco_brasil.csv", mode='a', header=False, index=False)
            
            # print('página lida: ', indice)
            # print(dfs)

            with open('Banco_brasil.csv', 'r', encoding="charmap") as file:
                reader = csv.reader(file)
                for row in reader:
                    lista.append(row)
            
        # Variáveis de valores:

        for row in lista:
                data = ''
                valores = ''
                descricao = ''
                mes = ''
                ano = ''
                primeiro_dia_mes = ''

                # Aqui estou limpando os dados da descrição, deixando ela limpa, apenas com informações do remetente/destinatário;
                
                for value in row:
                    if value:
                        descricao = value
                        break

                descricao = re.sub(retirando_desc, '', descricao)
                descricao = descricao.replace("\n", "")

                for value in row:
                    buscando_data = re.findall(data_pattern, value)
                    if buscando_data:
                        data = buscando_data[0]
                        mes = data[3:5]
                        ano = data[6:]
                        primeiro_dia_mes = f'01/{mes}/{ano}'

                for value in row:
                    buscando_valores = re.findall(value_pattern, value)
                    if buscando_valores:
                        valores = buscando_valores[0]
                        if '(-)' in valores:
                            
                            valores = valores.replace('(-)', '')
                            valores = float(f'-{valores.replace(".", "").replace(",", ".")}')
                        else:
                            valores = valores.replace('(+)', '')
                            valores = float(f'{valores.replace(".", "").replace(",", ".")}') 

                if data and valores and descricao:
                    row_value = Linha('Banco do Brasil', data, descricao, valores,'','', mes, ano, primeiro_dia_mes)
                    lista_formatada.append(row_value)

        os.remove('Banco_brasil.csv')
        return lista_formatada

    def lendo_pdf_brasil_v2(self, pdf):

        # variaveis de controle.
        num_pages = 0       
        indice = 0
        lista = []
        lista_formatada = []

        # Variáveis de Regex
        data_pattern = r'\b\d{2}/\d{2}/\d{4}\b'
        valor_pattern = r"\d{1,3}(?:\.\d{3})*,\d{2}\s[DC]"
        texto_pattern = r".*[a-zA-Z]+.*"

        # Variáveis para captura de dados
        data = ''
        primeiro_dia_mes = ''
        mes = ''
        ano = ''
        descricao = ''
        descricao_adicional = ''

        # Variáveis de controle.
        contador_ocorrencia = 0

        # Área das tabelas
        area_analisada = ''
        page_any = [33, 42, 840, 551]

        try:
            with open(pdf, 'rb') as pdf_file:
                pdf_reader = PyPDF2.PdfReader(pdf_file)
                num_pages = len(pdf_reader.pages)
        except:
            return print("número invalido")
        
        for i in range(num_pages):
            indice+=1

            # neste caso como a primeira página não exige de area, por tanto ela será localizada usando esse texto para aplicar a extração.
            texto_padrao = 'Consultas - Extrato de conta corrente'

            text = extraindo_texto_2(pdf, indice)

            if texto_padrao in text:
                dfs = tabula.read_pdf(pdf, pages=indice)
            else:
                area_analisada = page_any
                dfs = tabula.read_pdf(pdf, pages=indice, area=area_analisada)

            for table in dfs:
                table.to_csv("Banco_brasil_v2_csv", mode='a', index=False)

            # if indice == 9 or indice == 10:
            #     print(dfs)
            
        decodificando_csv(lista, 'Banco_brasil_v2_csv')

        os.remove('Banco_brasil_v2_csv')

        linha_anterior = lista[0]
        
        indice_row = 0
        linha_anterior = lista[0]

        for indice, row in enumerate(lista):

            contador = 0
            contador_unnamed = 0

            for value in row:
                if value:
                    contador+=1

            if indice >= 1:
                linha_anterior = lista[indice - 1]

            valores = 0
            descricao = ''
            descricao_adicional = ''

            for value in linha_anterior:
                validando_data = re.findall(data_pattern, value)
                if validando_data:
                    data = validando_data[0]
                    mes = data[3:5]
                    ano = data[6:]
                    primeiro_dia_mes = f'01/{mes}/{ano}'

            for value in linha_anterior:
                validando_valor = re.findall(valor_pattern, value)
                if validando_valor:
                    valores_limpar = validando_valor[0]

                    if 'D' in valores_limpar:
                        valores_limpar = valores_limpar.replace('D', '').replace(" ", "")
                        valores_limpar = valores_limpar.replace(".", "").replace(",", ".")
                        valores = float(f'-{valores_limpar}')
                    else:
                        valores_limpar = valores_limpar.replace('C', '').replace(" ", "")
                        valores_limpar = valores_limpar.replace(".", "").replace(",", ".")
                        valores = float(valores_limpar)

                    break
            
            for value in linha_anterior:
                validando_desc = re.findall(texto_pattern, value)
                if validando_desc:
                    # caso encontrar um valor inválido na descrição, ele irá ignorar.
                    if 'Unnamed' in validando_desc[0] and len(value) == 10 and contador > 1:
                        continue
                    else:
                        descricao = validando_desc[0]
                        break
            
            for value in row:
                if 'Unnamed' in value:
                    contador_unnamed+=1

            # Aqui ele irá pegar aquela row que tiver apenas um único valor verdadeiro.
            if contador == 1:
                for value in row:
                    validando_desc = re.findall(texto_pattern, value)
                    if validando_desc:
                        if 'Unnamed' in validando_desc[0]:
                            continue
                        else:
                            descricao_adicional = validando_desc[0]
                            descricao+=' '+descricao_adicional
                            break
            
            if contador_unnamed > 4:
                for value in row:
                    validando_desc = re.findall(texto_pattern, value)
                    if validando_desc:
                        if 'Unnamed:' in validando_desc[0]:
                            continue
                        else:
                            descricao_adicional = validando_desc[0]
                            descricao+=' '+descricao_adicional
                            break

            if valores and descricao:
                rowValue = Linha('Banco - Brasil - v2', data, descricao, valores,'','', mes, ano, primeiro_dia_mes)
                lista_formatada.append(rowValue)

        return lista_formatada

    def lendo_pdf_brasil_v3(self, pdf):

        # variaveis de controle.
        num_pages = 0       
        indice = 0
        area_analisada = ''

        # Área das tabelas
        area_analisada = ''
        page_main = [74, 40, 807, 522]
        page_any = [40, 35, 807, 522]

        # regex dos valores capturados.
        data_pattern = r"\d{2}/\d{2}/\d{4}"
        texto_pattern = r".*[a-zA-Z]+.*"
        valor_pattern = r"\d{1,3}(?:\.\d{3})*,\d{2}\s[DC]"

        #valores de dados
        primeiro_dia_mes = ''
        dia = ''
        mes = ''
        ano = ''

        # Armazenar valores
        lista = []
        lista_formatada = []

        try:
            with open(pdf, 'rb') as pdf_file:
                pdf_reader = PyPDF2.PdfReader(pdf_file)
                num_pages = len(pdf_reader.pages)
        except:
            return print("número invalido")

        for i in range(num_pages):
            indice+=1

            texto_padrao = 'SISBB - Sistema de Informações Banco do Brasil'

            text = extraindo_texto_2(pdf, indice)
            
            if texto_padrao in text:
                area_analisada = page_main 
                dfs = tabula.read_pdf(pdf, pages=indice, area=area_analisada)
            else:
                area_analisada = page_any
                dfs = tabula.read_pdf(pdf, pages=indice, area=area_analisada)

            # if indice == 1 or indice == 2:
            #     print(dfs)

            for table in dfs:
                table.to_csv("banco_brasil_v3_csv", mode='a', index=False)

        decodificando_csv(lista, 'banco_brasil_v3_csv')

        os.remove('banco_brasil_v3_csv')

        linha_anterior = lista[1]
        linha_ante_anterior = lista[0]

        for indice, row in enumerate(lista):

            valores = ''
            descricao = ''

            # contador de espaços vazios
            cont_empty_atual = 0
            cont_empty_ante_anterior = 0

            if indice >= 2:
                linha_anterior = lista[indice - 1]
                linha_ante_anterior = lista[indice - 2]
                # if indice == 3:
                #     print('linha analisado: ', linha_ante_anterior, linha_anterior, row)

            for value in row:
                if value == '':
                    cont_empty_atual+=1

            for value in linha_ante_anterior:
                if value == '':
                    cont_empty_ante_anterior+=1

            for value in linha_ante_anterior:
                buscando_data = re.findall(data_pattern, value)
                if buscando_data:
                    data = buscando_data[0]
                    mes = data[3:5]
                    ano = data[6:]
                    primeiro_dia_mes = f'01/{mes}/{ano}'
                    break

            for value in linha_ante_anterior:
                buscando_valores = re.findall(valor_pattern, value)
                if buscando_valores:
                    valores = buscando_valores[0]
                    if 'D' in valores:
                        valores = valores.replace('D', '')
                        valores = float(f'-{valores.replace(".", "").replace(",", ".")}')
                        break
                    else:
                        valores = valores.replace('C', '')
                        valores = float(f'{valores.replace(".", "").replace(",", ".")}') 
                        break
            
            for value in linha_ante_anterior:
                validando_desc = re.findall(texto_pattern, value)
                validando_desc_numero = re.findall(valor_pattern, value) # evitando que capture os valores númerico como descrição.
                if validando_desc and len(validando_desc_numero) == 0 and value != 'Unnamed: 0':
                    for desc in validando_desc:
                        descricao+=desc
                    break
            
            if cont_empty_ante_anterior == 5 and cont_empty_atual == 5:
                # print(valores, descricao)
                # print(linha_ante_anterior, linha_anterior, row)
                
                for value in linha_anterior:
                    buscando_data = re.findall(data_pattern, value)
                    if buscando_data:
                        data = buscando_data[0]
                        mes = data[3:5]
                        ano = data[6:]
                        primeiro_dia_mes = f'01/{mes}/{ano}'
                        break
                
                for value in linha_anterior:
                    buscando_valores = re.findall(valor_pattern, value)
                    if buscando_valores:
                        valores = buscando_valores[0]
                        if 'D' in valores:
                            valores = valores.replace('D', '')
                            valores = float(f'-{valores.replace(".", "").replace(",", ".")}')
                            break
                        else:
                            valores = valores.replace('C', '')
                            valores = float(f'{valores.replace(".", "").replace(",", ".")}') 
                            break
                
                for value in linha_ante_anterior:
                    if value and value != 'Unnamed: 0':
                        descricao=value

                for value in row:
                    if value and value != 'Unnamed: 0':
                        descricao+=' '+value
                
            if valores and descricao:
                rowValue = Linha('Banco - Brasil - v3', data, descricao, valores,'','', mes, ano, primeiro_dia_mes)
                lista_formatada.append(rowValue)
            
        return lista_formatada

# Leitor PDF Mercado pago;
class leitor_pdf_mercado_pago:

    def __init__(self):
        pass

    def leitor_pdf_mercado_pago_v1(self, pdf):
        
        # Variáveis de Indice:
        indice = 0

        # Variaveis para captura de dados
        lista = []
        areaPageUm = [[174, 22, 543, 421]]
        areaPageAll = [[33, 17, 558, 424]]
        area_analisada = areaPageAll
        lista_formatada = []

        # Regex para capturar dados:
        data_pattern = r'\b\d{2}-\d{2}-\d{4}\b'
        valor_pattern = r"^R\$ ?-?\d{1,3}(?:\.\d{3})*,\d{2}$"
        texto_pattern = r".*[a-zA-Z]+.*"

        # Variaveis de valores:
        data = ''
        valores = ''
        mes = ''
        ano = ''
        primeiro_dia_mes = ''
        texto_pattern = r".*[a-zA-Z]+.*"
        contador_ocorrencia = 0
        descricao_adicional = ''
        descricao = ''
        data_pattern = r'\b\d{2}-\d{2}-\d{4}\b'
        valor_pattern = r"^R\$ ?-?\d{1,3}(?:\.\d{3})*,\d{2}$"

        with open(pdf, 'rb') as pdf_file:
            pdf_reader = PyPDF2.PdfReader(pdf_file)
            num_pages = len(pdf_reader.pages)
        
        for i in range(num_pages):

            indice+=1

            reader = PyPDF2.PdfReader(pdf)
            page = reader.pages[i]
            text = page.extract_text()

            # print(f'pagina: {i}')

            # print(text)

            # print('------------x-------------')

            if 'EXTRATO DE CONTA' in text:
                area_analisada = areaPageUm
            else:
                area_analisada = areaPageAll
        
            dfs = tabula.read_pdf(pdf, pages=indice, area=area_analisada)

            for table in dfs:
                table.to_csv("mercado_pago_csv", mode='a', header=False, index=False)  

        decodificando_csv(lista, 'mercado_pago_csv')
        
        os.remove('mercado_pago_csv')

        for row in lista:

            # Pegando data
            for value in row:
                validando_data = re.findall(data_pattern, value)
                if validando_data:
                    data = validando_data[0]
                    mes = data[3:5]
                    ano = data[6:]
                    primeiro_dia_mes = f'01/{mes}/{ano}'
            
            for value in row:
                validando_valores = re.findall(valor_pattern, value)
                if validando_valores:
                    valores_limpar = validando_valores[0]
                    
                    # Arrumando valores pego na tabela, pois os mesmo estão em formato invalido.
                    if '-' in valores_limpar:

                        valores_limpar = valores_limpar.replace('R$', '')

                        valores_limpar = valores_limpar.replace('R$', '')
                        valores = valores_limpar
                        valores = float(valores.replace(".", "").replace(",", ".")) 

                        valores = float(valores)
                        break
                    else:
                        valores_limpar = valores_limpar.replace('R$', '')
                        valores = valores_limpar
                        valores = float(valores.replace(".", "").replace(",", "."))
                        break

            contador = 0

            # Contando quantos valores verdadeiros tem no array, se for igual a 1, significa que a descrição está picada;
            for value in row:
                if value:
                    contador+=1

            # Com a descrição picada, eu procuro o valor verdadeiro e realizo o incremento na variavel de descrição adicional, e incrementando no contador_ocorrencia;
            if contador == 1:
                for value in row:
                    if value:
                        descricao_adicional+=' '+value
                        contador_ocorrencia+=1

            # Caso o contador_ocorrencia for incrementado, ele irá rodar o loop mais uma vez para encontrar o segundo valor.
            if contador_ocorrencia == 1:
                continue
            
            # Porém se a descriçaõ vim completa, ele irá apenas procurar normalmente um padrão de texto.
            if contador_ocorrencia == 0 and contador > 1:
                for value in row:
                    validando_texto = re.findall(texto_pattern, value)
                    if validando_texto:
                        descricao = validando_texto[0]
                        break
            
            # Quando a descrição picada ficar completa, o contador_ocorrencia será dois, sendo assim adicionado na variavel descrição;
            if contador_ocorrencia == 2:
                descricao = descricao_adicional
                contador_ocorrencia = 0
                descricao_adicional = ''
            
            # A linha só será adicionada caso o contador de ocorrencia for zerado(Significa que ou ele achou e completou a descrição picada ou não foi necessário o processo.)
            # Aqui ele irá verificar de todosos valores foram preenchidos, e por fim validar se o último valor do indice é diferente de '', pois se sim, ele não é contabilizado
            # porque somente a sua descrição foi capturada.
            
            # Validando descrição, pois ela está pegando valores(reais) como descrição.
            validando_descricao = re.findall(valor_pattern, descricao)

            if contador_ocorrencia == 0 and data and valores and len(validando_descricao) < 1:
                rowValue = Linha('Mercado Pago', data, descricao, valores,'','', mes, ano, primeiro_dia_mes)
                lista_formatada.append(rowValue)
        
        return lista_formatada

# Leitor PDF PicPay;
class leitor_pdf_PicPay:

    def __init__(self):
        pass

    def leitor_pdf_PicPay_v1(self, pdf):

        # Variáveis de Indice:
        indice = 0

        # Variaveis para captura de dados
        lista = []
        areaPageUm = [[239, 6, 717, 583]]
        areaPageAll = [[33, 17, 558, 424]]
        area_analisada = areaPageUm
        lista_formatada = []

        # Regex para capturar dados:
        data_pattern = r"\b\d{2}/\d{2}/\d{4}\b"
        valor_pattern = r"-? ?R\$ ?\d{1,3}(?:\.\d{3})*,\d{2}"
        texto_pattern = r".*[a-zA-Z]+.*"

        # Variaveis de valores:
        data = ''
        mes = ''
        ano = ''
        primeiro_dia_mes = ''

        # Pegando o número de páginas:
        with open(pdf, 'rb') as pdf_file:
            pdf_reader = PyPDF2.PdfReader(pdf_file)
            num_pages = len(pdf_reader.pages)

        for i in range(num_pages):
            indice+=1

            dfs = tabula.read_pdf(pdf, pages=indice, area=area_analisada)

            for table in dfs:
                table.to_csv("PicPay_csv", mode='a', index=False)
                            
        with open('PicPay_csv', 'r') as file:
                reader = csv.reader(file)
                for row in reader:
                    lista.append(row) 

        os.remove('PicPay_csv')

        for row in lista:

            valores = ''
            descricao = ''
            
            # Capturando Data:
            for value in row:
                validando_data = re.findall(data_pattern, value)
                if validando_data:
                    data = validando_data[0]
                    mes = data[3:5]
                    ano = data[6:]
                    primeiro_dia_mes = f'01/{mes}/{ano}'

            # Capturando Valores:
            for value in row:
                validando_valores = re.findall(valor_pattern, value)
                if validando_valores:
                    valores_limpar = validando_valores[0]
                    
                    # Arrumando valores pego na tabela, pois os mesmo estão em formato invalido.
                    if '-' in valores_limpar:
                        # Formatando o valor para deixar no padrão do python para números de ponto flutuante.
                        valores_limpar = valores_limpar.replace('R$', '').replace('-', '').replace(".", "").replace(",", ".").replace(" ", "")

                        valores = float(f'-{valores_limpar}') 
                    else:
                        valores_limpar = valores_limpar.replace('R$', '')
                        valores = valores_limpar
                        valores = float(valores.replace(".", "").replace(",", ".")) 
                    break
            
            for value in row:
                validando_texto = re.findall(texto_pattern, value)
                if validando_texto:
                    descricao = validando_texto[0]
                    break

            if data and valores and descricao:
                rowValue = Linha('PicPay', data, descricao, valores,'','', mes, ano, primeiro_dia_mes)
                lista_formatada.append(rowValue)
        
        return lista_formatada

# Leitor PDF inter;
class leitor_pdf_inter:

    def __init__(self):
        pass

    def leitor_pdf_inter_v1(self, pdf):

        # Variaveis para armazenar dados
        lista = []
        lista_formatada = []

        # Regex para captura de valores:
        data_pattern = r'(\d{1,2})\s+de\s+(Janeiro|Fevereiro|Março|Abril|Maio|Junho|Julho|Agosto|Setembro|Outubro|Novembro|Dezembro)\s+de\s+(\d{4})'
        valores_pattern = r'(-?R\$\s*\d{1,3}(?:\.\d{3})*,\d{2})'
        texto_pattern = r'[A-Za-zÀ-ÖØ-öù-ÿ]+(?:\s+[A-Za-zÀ-ÖØ-öù-ÿ]+)*'

        # funções e vriáveis auxiliares
        meses = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho', 'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
        data = ''
        
        with open(pdf, 'rb') as pdf_file:
            pdf_reader = PyPDF2.PdfReader(pdf_file)
            num_pages = len(pdf_reader.pages)

        indice = 0

        for i in range(num_pages):
            indice+=1
            dfs = tabula.read_pdf(pdf, pages=indice, area=[180, 39, 743, 555])

            for table in dfs:
                table.to_csv("inter_csv", mode='a', index=False)
        

        with open('inter_csv', 'r') as file:
            reader = csv.reader(file)
            for row in reader:
                lista.append(row)
        
        os.remove('inter_csv')

        for row in lista:

            valores = ''
            descricao = ''

            # capturando data.
            for value in row:
                matches = re.findall(data_pattern, value)

                if matches:
                    trabalhando_data = matches[0]

                    dia = trabalhando_data[0]
                    mes = trabalhando_data[1]
                    ano = trabalhando_data[2]

                    # Aqui estou alterando o valor do mes para um valor númerico equivalente.
                    for i in range(len(meses)):
                        if mes == meses[i]:
                            if i+1 >= 10:
                                mes = i+1
                                break  
                            mes = f'0{i+1}'
                            break

                    # Agora aqui estou apenas adicionando um 0 a mais para valores abaixo de 9
                    dia = int(dia)
                    if dia <= 9:
                        dia = f'0{dia}'

                    primeiro_dia_mes = f'01/{mes}/{ano}'
                    data = f'{dia}/{mes}/{ano}'

            # capturando valores.
            for value in row:
                matches_valor = re.findall(valores_pattern, value)

                if matches_valor:
                    valores_limpar = matches_valor[0]

                    if '-' in valores_limpar:
                        valores_limpar = valores_limpar.replace('-R$', '').replace(" ", "")
                        valores_limpar = valores_limpar.replace(".", "").replace(",", ".")
                        valores = float(f'-{valores_limpar}')
                    else:
                        valores_limpar = valores_limpar.replace('R$', '').replace(" ", "")
                        valores_limpar = valores_limpar.replace(".", "").replace(",", ".")
                        valores = float(valores_limpar)
                    
            # capturando decrição.
            for value in row:
                matches_texto = re.findall(texto_pattern, value)
                matches_valor_texto = re.findall(valores_pattern, value)
                
                if matches_texto and matches_valor_texto:
                    descricao = f'{row[0]} {row[1]}'
                    break

            if valores:
                rowValue = Linha('Banco - inter', data, descricao, valores,'','', mes, ano, primeiro_dia_mes)
                lista_formatada.append(rowValue)
                descricao = ''
        
        return lista_formatada

    def leitor_pdf_inter_v2(self, pdf):

        # Regex para captura de dados
        data_pattern = r'(\d{1,2})\s+de\s+(Janeiro|Fevereiro|Março|Abril|Maio|Junho|Julho|Agosto|Setembro|Outubro|Novembro|Dezembro)\s+de\s+(\d{4})'
        valores_pattern = r'(-?R\$\s*\d{1,3}(?:\.\d{3})*,\d{2})'
        texto_pattern = r'[A-Za-zÀ-ÖØ-öù-ÿ]+(?:\s+[A-Za-zÀ-ÖØ-öù-ÿ]+)*'

        # variaveis de armazenamento de dados:
        lista = []
        lista_formatada = []

        # variaveis e funções auxiliares:
        meses = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho', 'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
        data = ''

        with open(pdf, 'rb') as file_path:
            reader = PyPDF2.PdfReader(file_path)
            num_pages = len(reader.pages)
        
        page_um = [194, 36, 783, 560]
        page_any = [38, 33, 780, 592]
    
        indice = 0

        for i in range(num_pages):
            
            indice+=1

            reader = PyPDF2.PdfReader(pdf)
            page = reader.pages[i]
            text = page.extract_text()

            if 'Instituição: Banco Inter' in text:
                area_analisada = page_um
            else:
                area_analisada = page_any

            dfs = tabula.read_pdf(pdf, pages=indice, area=area_analisada, stream=True)
        
            for table in dfs:
                table.to_csv('banco_inter_csv', mode='a', index=False)

        with open('banco_inter_csv', 'r') as file:
            reader = csv.reader(file)
            for row in reader:
                lista.append(row)
        
        os.remove('banco_inter_csv')

        for row in lista:
            # capturando a data
            for value in row:
                matches_data = re.findall(data_pattern, value)

                if matches_data:
                    trabalhando_data = matches_data[0]

                    dia = trabalhando_data[0]
                    mes = trabalhando_data[1]
                    ano = trabalhando_data[2]

                    # Aqui estou alterando o valor do mes para um valor númerico equivalente.
                    for i in range(len(meses)):
                        if mes == meses[i]:
                            if i+1 >= 10:
                                mes = i+1
                                break  
                            mes = f'0{i+1}'
                            break

                    # Agora aqui estou apenas adicionando um 0 a mais para valores abaixo de 9
                    dia = int(dia)
                    if dia <= 9:
                        dia = f'0{dia}'

                    primeiro_dia_mes = f'01/{mes}/{ano}'
                    data = f'{dia}/{mes}/{ano}'
            
            # capturando valores monetários
            for value in row:
                matches_valor = re.findall(valores_pattern, value)

                if matches_valor:
                    valores_limpar = matches_valor[0]

                    if '-' in valores_limpar:
                        valores_limpar = valores_limpar.replace('-R$', '').replace(" ", "")
                        valores_limpar = valores_limpar.replace(".", "").replace(",", ".")
                        valores = float(f'-{valores_limpar}')
                    else:
                        valores_limpar = valores_limpar.replace('R$', '').replace(" ", "")
                        valores_limpar = valores_limpar.replace(".", "").replace(",", ".")
                        valores = float(valores_limpar)

                    break
            
            # capturando descrições.
            for value in row:
                matches_texto = re.findall(texto_pattern, value)

                if matches_texto:
                    descricao = value
                    break
            
            if 'Saldo do dia' in descricao:
                continue

            if valores and descricao:
                rowValue = Linha('Banco - inter', data, descricao, valores,'','', mes, ano, primeiro_dia_mes)
                lista_formatada.append(rowValue)
        
        return lista_formatada

    def leitor_pdf_inter_v3(self, pdf):
        
        # Regex para captura de dados
        data_pattern = r'(\d{1,2})\s+de\s+(Janeiro|Fevereiro|Março|Abril|Maio|Junho|Julho|Agosto|Setembro|Outubro|Novembro|Dezembro)\s+de\s+(\d{4})'
        valores_pattern = r'(-?R\$\s*\d{1,3}(?:\.\d{3})*,\d{2})'
        texto_pattern = r'[A-Za-zÀ-ÖØ-öù-ÿ]+(?:\s+[A-Za-zÀ-ÖØ-öù-ÿ]+)*'

        # variaveis de armazenamento de dados:
        lista = []
        lista_formatada = []

        # variaveis e funções auxiliares:
        meses = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho', 'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
        data = ''

        with open(pdf, 'rb') as file_path:
            reader = PyPDF2.PdfReader(file_path)
            num_pages = len(reader.pages)
        
        page_um = [194, 36, 783, 560]
        page_any = [38, 33, 780, 592]
    
        indice = 0

        for i in range(num_pages):
            
            indice+=1

            reader = PyPDF2.PdfReader(pdf)
            page = reader.pages[i]
            text = page.extract_text()

            if 'Instituição: Banco Inter' in text:
                area_analisada = page_um
            else:
                area_analisada = page_any

            dfs = tabula.read_pdf(pdf, pages=indice, area=area_analisada, stream=True)
        
            for table in dfs:
                table.to_csv('banco_inter_csv', mode='a', index=False)

        with open('banco_inter_csv', 'r') as file:
            reader = csv.reader(file)
            for row in reader:
                lista.append(row)
        
        os.remove('banco_inter_csv')

        for row in lista:
            # capturando a data
            for value in row:
                matches_data = re.findall(data_pattern, value)

                if matches_data:
                    trabalhando_data = matches_data[0]

                    dia = trabalhando_data[0]
                    mes = trabalhando_data[1]
                    ano = trabalhando_data[2]

                    # Aqui estou alterando o valor do mes para um valor númerico equivalente.
                    for i in range(len(meses)):
                        if mes == meses[i]:
                            if i+1 >= 10:
                                mes = i+1
                                break  
                            mes = f'0{i+1}'
                            break

                    # Agora aqui estou apenas adicionando um 0 a mais para valores abaixo de 9
                    dia = int(dia)
                    if dia <= 9:
                        dia = f'0{dia}'

                    primeiro_dia_mes = f'01/{mes}/{ano}'
                    data = f'{dia}/{mes}/{ano}'
            
            # capturando valores monetários
            for value in row:
                matches_valor = re.findall(valores_pattern, value)

                if matches_valor:
                    valores_limpar = matches_valor[0]

                    if '-' in valores_limpar:
                        valores_limpar = valores_limpar.replace('-R$', '').replace(" ", "")
                        valores_limpar = valores_limpar.replace(".", "").replace(",", ".")
                        valores = float(f'-{valores_limpar}')
                    else:
                        valores_limpar = valores_limpar.replace('R$', '').replace(" ", "")
                        valores_limpar = valores_limpar.replace(".", "").replace(",", ".")
                        valores = float(valores_limpar)
                    break
            
            # capturando descrições.
            for value in row:
                matches_texto = re.findall(texto_pattern, value)

                if matches_texto:
                    descricao = value
                    break
            
            if 'Saldo do dia' in descricao:
                continue

            if valores and descricao:
                rowValue = Linha('Banco - inter', data, descricao, valores,'','', mes, ano, primeiro_dia_mes)
                lista_formatada.append(rowValue)
        
        return lista_formatada

# Leitor PDF sisprime;
class leitor_pdf_sisprime:

    def __init__(self):
        pass

    def leitor_pdf_sisprime_v1(self, pdf):

        # Regex para capturar dados:
        data_pattern = r'\b\d{2}/\d{2}/\d{4}\b'
        valor_pattern = r"^-?R\$ ?-?\d{1,3}(?:\.\d{3})*,\d{2}$"
        texto_pattern = r".*[a-zA-Z]+.*"

        # variaveis de armazenamento de dados:
        lista = []
        lista_formatada = []

        # Variaveis de valores:
        data = ''
        mes = ''
        ano = ''
        primeiro_dia_mes = ''

        # Capturando área
        page_um = [144, 0, 809, 576]
        page_any = [35, 0, 809, 576]

        # Pegando o número de páginas:
        with open(pdf, 'rb') as pdf_file:
            pdf_reader = PyPDF2.PdfReader(pdf_file)
            num_pages = len(pdf_reader.pages)
        
        # Variáveis de controle.
        indice = 0

        area_analisada = page_um

        for i in range(num_pages):
            indice+=1

            if indice == 1:
                area_analisada = page_um
            else:
                area_analisada = page_any
            
            dfs = tabula.read_pdf(pdf, pages=indice, area=area_analisada)
            
            for table in dfs:
                table.to_csv("Sisprime_csv", mode='a', index=False)
                
        decodificando_csv(lista, 'Sisprime_csv')

        os.remove('Sisprime_csv')

        for row in lista:

            # Variáveis de valores
            valores = 0
            descricao = ''

            # capturando data
            for value in row:
                matches_data = re.findall(data_pattern, value)

                if matches_data:
                    data = matches_data[0]
                    dia = data[:3]
                    mes = data[3:5]
                    ano = data[6:] 
                    primeiro_dia_mes = f'01/{mes}/{ano}'
            
            # capturando valores
            for value in row:

                matches_valores = re.findall(valor_pattern, value)

                if matches_valores:
                    valores_limpar = matches_valores[0]
                    
                    # aqui estou tendo problemas para identificar quem é o valor negativo, por tanto no padrão quem for positivo os últimos dois
                    # indice estarão preenchidos com valores que correpondem ao padrão.
                    matches_ult_valor = ''
                    matches_peUlt_valor = ''

                    # validando ultimos dois valores com regex de números em formato monetário.
                    if len(row) > 2:
                        matches_ult_valor = re.findall(valor_pattern, row[-1])
                        matches_peUlt_valor = re.findall(valor_pattern, row[-2])

                    if matches_peUlt_valor and matches_ult_valor:
                        valores_limpar = valores_limpar.replace('R$', '').replace(" ", "")
                        valores_limpar = valores_limpar.replace(".", "").replace(",", ".")
                        valores = float(valores_limpar)
                    else:
                        valores_limpar = valores_limpar.replace('R$', '').replace(" ", "")
                        valores_limpar = valores_limpar.replace(".", "").replace(",", ".")
                        valores = float(f'-{valores_limpar}')
                    break

            # capturando descrição:
            for value in row:
                matches_texto = re.findall(texto_pattern, value)
                matches_valores_desc = re.findall(valor_pattern, value)

                if matches_texto and len(matches_valores_desc) == 0:
                    descricao+=' '+value
                
            if valores and descricao:
                rowValue = Linha('Banco - Sisprime', data, descricao, valores,'','', mes, ano, primeiro_dia_mes)
                lista_formatada.append(rowValue)
        
        return lista_formatada

# leitor PDF sicredi;
class leitor_pdf_sicredi:

    def __init__(self):
        pass

    def leitor_pdf_sicredi_v1(self, pdf):
       # variaveis de controle.
        num_pages = 0       
        indice = 0

        # Área das tabelas
        area_analisada = ''
        page_um = [345,22,819,568]
        page_any = [27,25,820,565]

        # regex dos valores capturados.
        data_pattern = r"\d{2}/\d{2}/\d{4}"
        texto = r".*[a-zA-Z]+.*"
        validando_valor = r"[+-]\s*R\$\s*\d{1,3}(?:\.\d{3})*,\d{2}"

        #valores de dados
        primeiro_dia_mes = ''
        dia = ''
        mes = ''
        ano = ''
        descricao = ''

        # Armazenar valores
        lista = []
        lista_formatada = []

        try:
            with open(pdf, 'rb') as pdf_file:
                pdf_reader = PyPDF2.PdfReader(pdf_file)
                num_pages = len(pdf_reader.pages)
        except:
            return print("número invalido")
        
        # dfs = tabula.read_pdf(pdf, pages="1", area=page_um)

        # print(dfs)

        for i in range(num_pages):
            indice+=1

            texto_padrao = 'Extrato de conta corrente'

            text = extraindo_texto_2(pdf, indice)

            if texto_padrao in text:
                area_analisada = page_um
            else:
                area_analisada = page_any

            
            dfs = tabula.read_pdf(pdf, pages=indice, area=area_analisada)

            for table in dfs:
                table.to_csv("Sicred_v1_csv", mode='a', header=False, index=False)

        decodificando_csv(lista, 'Sicred_v1_csv')

        os.remove('Sicred_v1_csv')

        for row in lista:
            
            data = ''
            valores = 0

            for value in row:
                buscando_data = re.findall(data_pattern, value)
                if buscando_data:
                    data = buscando_data[0]
                    mes = data[3:5]
                    ano = data[6:]
                    primeiro_dia_mes = f'01/{mes}/{ano}'
            
            for value in row:
                    buscando_valores = re.findall(validando_valor, value)
                    if buscando_valores:
                        valores = buscando_valores[0]
                        if '-' in valores:
                            valores = valores.replace(' ', '').replace('-', '').replace('R$', '')
                            valores = float(f'-{valores.replace(".", "").replace(",", ".")}')
                        else:
                            valores = valores.replace(' ', '').replace('+', '').replace('R$', '')
                            valores = float(f'{valores.replace(".", "").replace(",", ".")}') 
            
            for value in row:
                buscando_texto = re.findall(texto, value)
                if buscando_texto:
                    descricao = buscando_texto[0]
                    break
            
            if valores != 0:
                rowValue = Linha('Banco - Sicredi - v1', data, descricao, valores,'','', mes, ano, primeiro_dia_mes)
                lista_formatada.append(rowValue)
            
        return lista_formatada

    def leitor_pdf_sicredi_v2(self, pdf):
        # variaveis de controle.
        num_pages = 0       
        indice = 0

        # Área das tabelas
        area_analisada = ''
        page_um = [146, 40, 803, 558]
        page_any = [42, 31, 805, 555]

        # regex dos valores capturados.
        data_pattern = r"\d{2}/\d{2}/\d{4}"
        texto = r".*[a-zA-Z]+.*"
        validando_valor = r"-?\d{1,3}(?:\.\d{3})*,\d{2}"

        #valores de dados
        primeiro_dia_mes = ''
        dia = ''
        mes = ''
        ano = ''
        descricao = ''
        data = ''

        # Armazenar valores
        lista = []
        lista_formatada = []

        try:
            with open(pdf, 'rb') as pdf_file:
                pdf_reader = PyPDF2.PdfReader(pdf_file)
                num_pages = len(pdf_reader.pages)
        except:
            return print("número invalido")
        
        texto_padrao = 'Data Descrição Documento Valor (R$) Saldo (R$)'

        for i in range(num_pages):
            indice+=1

            text = extraindo_texto_2(pdf, indice)

            if texto_padrao in text:
                area_analisada = page_um
                dfs = tabula.read_pdf(pdf, pages=indice, area=area_analisada)
            else:
                dfs = tabula.read_pdf(pdf, pages=indice)

            print(dfs)

            for table in dfs:
                table.to_csv("sicredi_csv", mode='a', index=False)

        decodificando_csv(lista, 'sicredi_csv')

        os.remove('sicredi_csv')

        for row in lista:

            valores = 0

            for value in row:
                buscando_data = re.findall(data_pattern, value)
                if buscando_data:
                    data = buscando_data[0]
                    mes = data[3:5]
                    ano = data[6:]
                    primeiro_dia_mes = f'01/{mes}/{ano}'

            for value in row:
                    buscando_valores = re.findall(validando_valor, value)
                    if buscando_valores:
                        valores = buscando_valores[0]
                        if '-' in valores:
                            valores = valores.replace(' ', '').replace('-', '')
                            valores = float(f'-{valores.replace(".", "").replace(",", ".")}')
                        else:
                            valores = valores.replace(' ', '').replace('+', '')
                            valores = float(f'{valores.replace(".", "").replace(",", ".")}')
                        break

            for value in row:
                buscando_texto = re.findall(texto, value)
                if buscando_texto:
                    descricao = buscando_texto[0]
                    break

            if descricao and valores:
                rowValue = Linha('Banco - Sicredi - v2', data, descricao, valores,'','', mes, ano, primeiro_dia_mes)
                lista_formatada.append(rowValue)
        
        return lista_formatada           

    def leitor_pdf_sicredi_v3(self, pdf):
        #valores de dados
        primeiro_dia_mes = ''
        mes = ''
        ano = ''
        data = ''

        # armazenamento de dados
        lista_formatada = []
        lista_dados = []

        instanciando_classe = Estrutura_Padrao(pdf, [116, 71, 816, 533], [70, 24, 606, 534], 'Associado:', 'banco_sicoob_csv', lista_dados, 2, False)

        instanciando_classe.main_reader()

        for indice, row in enumerate(lista_dados):

            descricao = ''
            valores = 0

            for value in lista_dados[indice]:
                match_data = Estrutura_Padrao.regex_padrao_data(value, 2)
                if match_data:
                    data = match_data[0]
                    mes = data[3:5]
                    ano = data[6:]
                    primeiro_dia_mes = f'01/{mes}/{ano}'
            
            for value in lista_dados[indice]:
                match_valor = Estrutura_Padrao.regex_padrao_valores(value, 1)
                if match_valor:
                    valores = match_valor[0]
                    valores = Estrutura_Padrao.formantando_numeros(valores, 1)
                    break
            
            for value in lista_dados[indice]:
                match_descricao = Estrutura_Padrao.regex_padrao_texto(value, 1)
                if match_descricao:
                    if 'Unnamed' in match_descricao[0]:
                        continue
                    else:
                        descricao = match_descricao[0]
                        break
            
            if valores != 0 and descricao:
                rowValue = Linha('banco Sicredi - V3', data, descricao, valores,'', '', mes, ano, primeiro_dia_mes)
                lista_formatada.append(rowValue)
        
        return lista_formatada
    
# leitor PDF C6;
class leitor_pdf_cSix:

    def __init__(self):
        pass
    
    def leitor_pdf_CSIX_v1(self, pdf):
        # variaveis de controle.
        num_pages = 0       
        indice = 0

        # Área das tabelas
        area_analisada = ''
        page_um = [155,30,811,566]
        page_any = [29,27,813,563]

        # regex dos valores capturados.
        data_pattern = r"\d{2}/\d{2}"
        texto = r".*[a-zA-Z]+.*"
        validando_valor = r'R\$ \d{1,3}(\.\d{3})*,\d{2}'
        teste_valor = r"-?\d{1,3}(?:\.\d{3})*,\d{2}"

        #valores de dados
        mes = ''
        ano = ''
        primeiro_dia_mes = ''
        data = ''

        # Armazenar valores
        lista = []
        lista_formatada = []
    
        # variaveis aulixiares 
        texto_padrao = 'Extrato exportado no dia'

        try:
            with open(pdf, 'rb') as pdf_file:
                pdf_reader = PyPDF2.PdfReader(pdf_file)
                num_pages = len(pdf_reader.pages)
        except:
            return print("número invalido")
        
        for i in range(num_pages):
            indice+=1

            # print(f'Página lida: {indice}')

            text = extraindo_texto_2(pdf, indice)

            if texto_padrao in text:
                area_analisada = page_um
            else:
                area_analisada = page_any
                
            dfs = tabula.read_pdf(pdf, pages=indice, stream=True, area=area_analisada)

            for table in dfs:
                table.to_csv("c6_v1_csv", mode='a', header=False, index=False)

        decodificando_csv(lista, 'c6_v1_csv')

        os.remove('c6_v1_csv')

        for row in lista:    

            descricao = ''
            valores = ''
            
            flag_type_value = True
            
            if ('Saída PIX' in row[0] or 'Saída PIX' in row[1]) or ('Pagamento' in row[0] or 'Pagamento' in row[1]) or ('Outros gastos' in row[0] or 'Outros gastos' in row[1]):
                flag_type_value = False

            # capturando data
            for value in row:
                validando_data = re.match(data_pattern, value)
                if validando_data:
                    
                    data = validando_data[0]
                    mes = data[3:5]

                    # Alguns extratos não tem o ano ao lado do mês e dia da transferência, então para ajustar isso apenas valido se o Mês é igual ou menor ao mês capturado, caso sim, se trata de um extrato de 2024, se não é um extrato do ano anterior.
                    mes_atual = data_atual.month

                    if(int(mes) <= mes_atual):
                        ano = data_atual.year
                    else:
                        ano = data_atual.year - 1

                    mes = data[3:5]
                    data = f'{validando_data[0]}/{ano}'

                    primeiro_dia_mes = f'01/{mes}/{ano}'
                    break

            for value in row:
                match = re.match(validando_valor, value)

                if match:
                    valores = value
                    if flag_type_value is False:
                        valores = valores.replace('R$', '').replace(' ', '')
                        valores = valores.replace(".","").replace(",", '.')
                        valores = float(f'-{valores}')
                    else:
                        valores = valores.replace('R$', '').replace(' ', '')
                        valores = valores.replace(".","").replace(",", '.')
                        valores = float(f'{valores}')
                    break
            
            try:
                flag_desc = True

                if row[2] == '':
                    match_texto = re.match(texto, row[1])
                    flag_desc = False
                else:
                    match_texto = re.match(texto, row[2])

                if match_texto and flag_desc is True:
                    descricao = row[2]
                else:
                    descricao = row[1]
            except:
                print(f'Erro na captura da descrição')

            if valores and descricao:
                # print('Banco - C6', data, descricao, valores,'','', mes, ano, primeiro_dia_mes)
                rowValue = Linha('Banco - C6', data, descricao, valores,'','', mes, ano, primeiro_dia_mes)
                lista_formatada.append(rowValue)

        return lista_formatada

# leitor PDF sicoob:
class leitor_pdf_SICOOB:

    def __init__(self):
        pass

    def lendo_pdf_banco_sicoob(self, pdf):
        #valores de dados
        primeiro_dia_mes = ''
        mes = ''
        ano = ''
        data = ''

        # armazenamento de dados
        lista_formatada = []
        lista_dados = []
        
        instanciando_classe = Estrutura_Padrao(pdf, [125, 108, 826, 486], [105, 26, 825, 482], 'SISTEMA DE COOPERATIVAS', 'banco_sicoob_extratoFeio_csv', lista_dados, 2, False)

        instanciando_classe.main_reader()

        for row in lista_dados:
            print(row)


class leitor_pdf_next:
    
    def __init__(self):
        pass
    
    # problemas na leitura da descrição.
    def leitor_pdf_next_v1(self, pdf):
        #valores de dados
        primeiro_dia_mes = ''
        mes = ''
        ano = ''
        data = ''

        # armazenamento de dados
        lista_formatada = []
        lista_dados = []

        instanciando_classe = Estrutura_Padrao(pdf, [210, 28, 791, 572], [107, 93, 772, 557], 'Banco de origem:', 'banco_next_csv', lista_dados, 2, False)

        instanciando_classe.main_reader()

        linha_anterior = ''
        linha_ante_anterior = ''

        for indice, row in enumerate(lista_dados):

            valores = 0

            if indice >= 2:
                linha_anterior = lista_dados[indice - 1]
                linha_ante_anterior = lista_dados[indice - 2]

            descricao = Estrutura_Padrao.leitura_desc_separado(lista_dados[indice], linha_anterior, linha_ante_anterior, 1)

            for value in linha_anterior:
                match_data = Estrutura_Padrao.regex_padrao_data(value, 2)
                if match_data:
                    data = match_data[0]
                    mes = data[3:5]
                    ano = data[6:]
                    primeiro_dia_mes = f'01/{mes}/{ano}'
            
            for value in linha_anterior:
                match_valor = Estrutura_Padrao.regex_padrao_valores(value, 3)
                if match_valor:
                    valores = match_valor[0]
                    valores = Estrutura_Padrao.formantando_numeros(valores, 2)
                    break
            
            if valores != 0 and descricao:
                rowValue = Linha('banco netx - V1', data, descricao, valores,'', '', mes, ano, primeiro_dia_mes)
                lista_formatada.append(rowValue)
        
        return lista_formatada

class leitor_pdf_sofisa:
    
    def __init__(self):
        pass
    
    def leitor_pdf_sofisa_v1(self, pdf):
        #valores de dados
        primeiro_dia_mes = ''
        mes = ''
        ano = ''
        data = ''

        # armazenamento de dados
        lista_formatada = []
        lista_dados = []

        instanciando_classe = Estrutura_Padrao(pdf, [262, 38, 799, 554], [32, 32, 826, 555], 'Extrato do período', 'banco_sofisa_csv', lista_dados, 2, False)

        instanciando_classe.main_reader()

        linha_anterior = ''
        linha_ante_anterior = ''

        for indice, row in enumerate(lista_dados):
           
            valores = 0
            descricao = ''

            for value in row:
                buscando_data = Estrutura_Padrao.regex_padrao_data(value, 2)
                if buscando_data:
                    data = buscando_data[0]
                    mes = data[3:5]
                    ano = data[6:]
                    primeiro_dia_mes = f'01/{mes}/{ano}'
                    break

            for value in row:
                    match_descricao = regex_padrao_texto(value, 1)
                    if match_descricao:
                        if 'Unnamed' in match_descricao[0]:
                            continue
                        else:
                            descricao = match_descricao[0]
                            break
            
            for value in row:
                match_valor = regex_padrao_valores(value, 1)
                if match_valor:
                    valores = match_valor[0]
                    valores = Estrutura_Padrao.formantando_numeros(valores, 1)
                    break
        
            if valores and descricao:
                # print('Banco - C6', data, descricao, valores,'','', mes, ano, primeiro_dia_mes)
                rowValue = Linha('Banco - Sofisa', data, descricao, valores,'','', mes, ano, primeiro_dia_mes)
                lista_formatada.append(rowValue)

        return lista_formatada

class PdfReaderVersion:
    
    def __init__(self):
        pass
    
    def extraindo_texto(self, pdf):
        try:
            reader = PyPDF2.PdfReader(pdf)
            page = reader.pages[0]
            text = page.extract_text()
        except:
            text = extract_text(pdf)

        return text

class to_excel:

    def localizando_substiruindo(self, arquivo_excel):
        wb = openpyxl.load_workbook(arquivo_excel)
        sheet = wb.active

        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and '.' in cell.value:
                    # Substituir ponto por vírgula nas células de valores numéricos
                    cell.value = cell.value.replace('.', ',')
        
        wb.save(arquivo_excel)

    def transformando_excel(self, lista, codificacao):

        with open('extrato_lido.csv', 'w', newline='', errors='replace') as arquivo_csv:
            escritor_csv = csv.writer(arquivo_csv)

            escritor_csv.writerow(["banco", "data", "desc", 'Movimentacao', "valor", 'saldo_extrato', 'saldo_calculado', 'mes', 'ano', 'primeiro_dia_mes'])

            for row in lista:
                escritor_csv.writerow([row.banco, row.data, row.descricao, row.descricao, row.valores, row.saldo_extrato, row.saldo_calculado, row.mes, row.ano, row.primeiro_dia_mes])

        try:
            df = pd.read_csv('extrato_lido.csv', encoding='mac_roman')
        except:
            return print("Formatação incorreta")

        # Aplicar a formatação para exibir com vírgula como separador decimal
    
        os.remove('extrato_lido.csv')
        df.to_excel('Extratos.xlsx', index=False, sheet_name="Extratos")

        # infelizmente a formatação float do python não funciona em alguns casos dentro do código, neste caso é preciso aplicar o formato novamente.
        arquivo = load_workbook('Extratos.xlsx') # abrindo o excel.

        aba_extrato = arquivo['Extratos'] # pegando a planilha correta.

        i=1 # o loop é iniciado em 1 pois aqui tenho de descartar a legenda que está nesta posição
        while True:
            i += 1 # logo em seguida ele é incrementado.
            valor_coluna = aba_extrato.cell(row=i, column=5).value # o array percorre e pega a biblioteca.

            if valor_coluna == None: # caso a célula for vazia, o valor None será atribuido a ela, por tanto significa que os valores chegaram a última recorrência.
                arquivo.save('Extratos.xlsx') # extrato salvo com a nova formatação.
                break
            
            verificando_tipo = type(valor_coluna) # pegando o tipo para validar se o erro se aplica a este extrato.

            # se não ele apenas continua o loop.
            if verificando_tipo is int:
                continue
            if verificando_tipo is str: # caso for str, ele irá realizar a formatação.
                
                if ',' in valor_coluna: # alguns vem com ',' ainda, po tanto reforço para tirar o valor invalido.
                    valor_coluna = valor_coluna.replace('.', '').replace(',', '.') # replace dos valores.
                    valor_coluna = float(valor_coluna) # formatando em float.
                else:
                    valor_coluna = float(valor_coluna) # formatando em float.

                aba_extrato.cell(row=i, column=5).value = valor_coluna # atribuindo o valor a coluna.

                aba_extrato.cell(row=i, column=5).number_format = '0.00' # bom, funcionou e esse cara tava aqui, ou seja, não sei se está ajudando, porém não se mexe em time que tá ganhando.